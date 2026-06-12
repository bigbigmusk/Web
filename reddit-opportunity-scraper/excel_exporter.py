"""把识别结果导出为格式化的 Excel 文件。"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# (表头, 取值键, 列宽)
COLUMNS = [
    ("子版块", "subreddit", 14),
    ("标题", "title", 50),
    ("是否机会", "is_opportunity_cn", 10),
    ("AI置信度", "ai_confidence", 9),
    ("规则分", "rule_score", 8),
    ("机会类型", "opportunity_type", 14),
    ("目标用户", "target_audience", 22),
    ("核心痛点", "pain_point", 32),
    ("建议方案", "suggested_solution", 32),
    ("变现方式", "monetization", 14),
    ("判断理由", "reasoning", 40),
    ("命中信号", "matched_signals_str", 24),
    ("点赞", "score", 7),
    ("评论数", "num_comments", 8),
    ("作者", "author", 16),
    ("发布时间", "created_str", 18),
    ("链接", "permalink", 40),
]

_HEADER_FILL = PatternFill("solid", fgColor="2D3142")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_OPP_FILL = PatternFill("solid", fgColor="D8F3DC")  # 是机会：淡绿
_WRAP = Alignment(wrap_text=True, vertical="top")


def _row_values(rec: dict) -> dict:
    """把内部记录补齐成导出所需的展示字段。"""
    created = rec.get("created_utc", 0) or 0
    created_str = (
        dt.datetime.utcfromtimestamp(created).strftime("%Y-%m-%d %H:%M")
        if created
        else ""
    )
    signals = rec.get("matched_signals") or []
    return {
        **rec,
        "is_opportunity_cn": "✅ 是" if rec.get("is_opportunity") else "—",
        "matched_signals_str": ", ".join(signals),
        "created_str": created_str,
    }


def export(records: list[dict], output_dir: str, filename: str | None = None) -> str:
    """导出记录到 xlsx，返回文件完整路径。

    records 已按重要度（是否机会、置信度、规则分）排序更佳，但本函数也会再排一次。
    """
    os.makedirs(output_dir, exist_ok=True)
    if not filename:
        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reddit_opportunities_{ts}.xlsx"
    path = os.path.join(output_dir, filename)

    # 排序：是机会优先 → AI 置信度高优先 → 规则分高优先
    records = sorted(
        records,
        key=lambda r: (
            r.get("is_opportunity", False),
            r.get("ai_confidence", 0),
            r.get("rule_score", 0),
        ),
        reverse=True,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "商业机会"

    # 表头
    for col_idx, (header, _key, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    # 数据行
    for r_idx, rec in enumerate(records, start=2):
        row = _row_values(rec)
        is_opp = rec.get("is_opportunity")
        for c_idx, (_header, key, _w) in enumerate(COLUMNS, start=1):
            value = row.get(key, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = _WRAP
            if is_opp:
                cell.fill = _OPP_FILL
            if key == "permalink" and value:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")

    wb.save(path)
    return path
